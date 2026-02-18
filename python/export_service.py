"""
Export Service - Generate Excel and PDF reports.
"""

import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False


class ExportService:
    def __init__(self, db):
        self.db = db
        # Default export directory
        self.export_dir = Path.home() / 'Documents' / 'BiometricSync Reports'
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(self, params):
        """Export attendance data to Excel."""
        if not HAS_OPENPYXL:
            raise Exception('Excel export not available. Install openpyxl: pip install openpyxl')

        date_from = params.get('dateFrom')
        date_to = params.get('dateTo')
        device_id = params.get('deviceId')
        report_type = params.get('reportType', 'detailed')

        # Get data
        result = self.db.get_attendance_logs(
            page=1,
            limit=10000,  # Get all records
            device_id=device_id if device_id else None,
            date_from=date_from,
            date_to=date_to
        )
        logs = result.get('logs', [])

        # Get devices for lookup
        devices = {d['id']: d['name'] for d in self.db.get_devices()}

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance Report'

        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        if report_type == 'detailed':
            # Detailed report - all records
            headers = ['S.No', 'User ID', 'Date', 'Time', 'Type', 'Device', 'Sync Status', 'Synced At']
            ws.append(headers)

            # Style headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Add data
            for idx, log in enumerate(logs, 1):
                timestamp = datetime.fromisoformat(log['timestamp']) if log['timestamp'] else None
                synced_at = datetime.fromisoformat(log['synced_at']) if log.get('synced_at') else None

                row = [
                    idx,
                    log['user_id'],
                    timestamp.strftime('%Y-%m-%d') if timestamp else '',
                    timestamp.strftime('%H:%M:%S') if timestamp else '',
                    'Check In' if log.get('punch_type', 0) == 0 else 'Check Out',
                    devices.get(log.get('device_id'), 'Unknown'),
                    'Synced' if log.get('synced') else 'Pending',
                    synced_at.strftime('%Y-%m-%d %H:%M:%S') if synced_at else ''
                ]
                ws.append(row)

                # Apply border to data cells
                for col in range(1, len(headers) + 1):
                    ws.cell(row=idx + 1, column=col).border = border

            # Set column widths
            column_widths = [8, 15, 12, 10, 12, 20, 12, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

        else:
            # Summary report - aggregated by user
            summary = {}
            for log in logs:
                user_id = log['user_id']
                if user_id not in summary:
                    summary[user_id] = {
                        'check_ins': 0,
                        'check_outs': 0,
                        'first_punch': None,
                        'last_punch': None
                    }

                timestamp = datetime.fromisoformat(log['timestamp']) if log['timestamp'] else None

                if log.get('punch_type', 0) == 0:
                    summary[user_id]['check_ins'] += 1
                else:
                    summary[user_id]['check_outs'] += 1

                if timestamp:
                    if summary[user_id]['first_punch'] is None or timestamp < summary[user_id]['first_punch']:
                        summary[user_id]['first_punch'] = timestamp
                    if summary[user_id]['last_punch'] is None or timestamp > summary[user_id]['last_punch']:
                        summary[user_id]['last_punch'] = timestamp

            headers = ['S.No', 'User ID', 'Total Check-Ins', 'Total Check-Outs', 'First Punch', 'Last Punch']
            ws.append(headers)

            # Style headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Add data
            for idx, (user_id, data) in enumerate(sorted(summary.items()), 1):
                row = [
                    idx,
                    user_id,
                    data['check_ins'],
                    data['check_outs'],
                    data['first_punch'].strftime('%Y-%m-%d %H:%M') if data['first_punch'] else '',
                    data['last_punch'].strftime('%Y-%m-%d %H:%M') if data['last_punch'] else ''
                ]
                ws.append(row)

                for col in range(1, len(headers) + 1):
                    ws.cell(row=idx + 1, column=col).border = border

            # Set column widths
            column_widths = [8, 15, 15, 15, 18, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

        # Save file
        filename = f'attendance_report_{date_from}_to_{date_to}.xlsx'
        filepath = self.export_dir / filename
        wb.save(str(filepath))

        return {
            'success': True,
            'path': str(filepath),
            'records': len(logs)
        }

    def export_to_pdf(self, params):
        """Export attendance data to PDF."""
        if not HAS_REPORTLAB:
            raise Exception('PDF export not available. Install reportlab: pip install reportlab')

        date_from = params.get('dateFrom')
        date_to = params.get('dateTo')
        device_id = params.get('deviceId')
        report_type = params.get('reportType', 'detailed')

        # Get data
        result = self.db.get_attendance_logs(
            page=1,
            limit=10000,
            device_id=device_id if device_id else None,
            date_from=date_from,
            date_to=date_to
        )
        logs = result.get('logs', [])

        # Get devices for lookup
        devices = {d['id']: d['name'] for d in self.db.get_devices()}

        # Create PDF
        filename = f'attendance_report_{date_from}_to_{date_to}.pdf'
        filepath = self.export_dir / filename

        doc = SimpleDocTemplate(
            str(filepath),
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,  # Center
            spaceAfter=20
        )
        elements.append(Paragraph('Attendance Report', title_style))

        # Subtitle with date range
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=1,
            spaceAfter=20
        )
        elements.append(Paragraph(f'Period: {date_from} to {date_to}', subtitle_style))
        elements.append(Spacer(1, 20))

        # Table style
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ])

        if report_type == 'detailed':
            headers = ['#', 'User ID', 'Date', 'Time', 'Type', 'Device', 'Status']
            data = [headers]

            for idx, log in enumerate(logs[:500], 1):  # Limit to 500 for PDF
                timestamp = datetime.fromisoformat(log['timestamp']) if log['timestamp'] else None
                row = [
                    str(idx),
                    log['user_id'],
                    timestamp.strftime('%Y-%m-%d') if timestamp else '',
                    timestamp.strftime('%H:%M:%S') if timestamp else '',
                    'In' if log.get('punch_type', 0) == 0 else 'Out',
                    devices.get(log.get('device_id'), 'Unknown')[:15],
                    'Synced' if log.get('synced') else 'Pending'
                ]
                data.append(row)

            col_widths = [30, 80, 80, 60, 50, 120, 60]

        else:
            # Summary
            summary = {}
            for log in logs:
                user_id = log['user_id']
                if user_id not in summary:
                    summary[user_id] = {'check_ins': 0, 'check_outs': 0}

                if log.get('punch_type', 0) == 0:
                    summary[user_id]['check_ins'] += 1
                else:
                    summary[user_id]['check_outs'] += 1

            headers = ['#', 'User ID', 'Check-Ins', 'Check-Outs', 'Total']
            data = [headers]

            for idx, (user_id, stats) in enumerate(sorted(summary.items()), 1):
                total = stats['check_ins'] + stats['check_outs']
                row = [
                    str(idx),
                    user_id,
                    str(stats['check_ins']),
                    str(stats['check_outs']),
                    str(total)
                ]
                data.append(row)

            col_widths = [40, 120, 100, 100, 100]

        table = Table(data, colWidths=col_widths)
        table.setStyle(table_style)
        elements.append(table)

        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray
        )
        elements.append(Paragraph(
            f'Generated on {datetime.now().strftime("%Y-%m-%d %H:%M:%S")} | Total Records: {len(logs)}',
            footer_style
        ))

        doc.build(elements)

        return {
            'success': True,
            'path': str(filepath),
            'records': len(logs)
        }
